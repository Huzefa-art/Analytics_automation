import requests
import re
import time
import sys
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Suppress SSL warnings ──────────────────────────────────────────────────────
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS_REQ = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ─────────────────────────────────────────────────────────────────────────────
# FINGERPRINTS  { category: { tech_name: [patterns] } }
# patterns are checked against: html_lower, response headers, cookies, scripts
# ─────────────────────────────────────────────────────────────────────────────
FINGERPRINTS = {
    "CMS": {
        "WordPress":     ["wp-content", "wp-includes", "wp-json", "xmlrpc.php", "/wp-login"],
        "Shopify":       ["cdn.shopify.com", "shopify.theme", "myshopify.com", "shopify/assets"],
        "Wix":           ["wix.com/", "wixsite.com", "parastorage.com", "wixstatic.com"],
        "Squarespace":   ["squarespace.com", "static1.squarespace", "sqsp.net"],
        "Webflow":       ["webflow.com", "assets.website-files.com", "uploads-ssl.webflow"],
        "Drupal":        ["drupal.settings", "/sites/default/files", "drupal.js", "x-generator: drupal"],
        "Joomla":        ["/media/jui/", "joomla", "/components/com_"],
        "Magento":       ["mage.cookies", "/skin/frontend/", "magento", "mage/"],
        "BigCommerce":   ["cdn.bigcommerce.com", "bigcommerce", "stencil"],
        "Ghost":         ["ghost.io", "content/themes", "ghost/"],
        "Hubspot CMS":   ["hs-sites.com", "hubspotpreview.com"],
        "PrestaShop":    ["prestashop", "/themes/classic/", "presta_shop"],
        "OpenCart":      ["opencart", "catalog/view/theme"],
        "WooCommerce":   ["woocommerce", "wc-ajax", "wc_add_to_cart"],
        "Webnode":       ["webnode.com"],
        "Weebly":        ["weebly.com", "weeblycloud.com"],
        "Blogger":       ["blogger.com", "blogspot.com", "www.gstatic.com/blogger"],
    },

    "CRM / Marketing Automation": {
        "HubSpot":          ["hs-scripts.com", "hubspot.com", "_hsp", "hsforms.com", "hs-analytics"],
        "Salesforce":       ["salesforce.com", "pardot.com", "force.com", "exacttarget.com"],
        "Zoho CRM":         ["zoho.com", "zohopublic.com", "zohocrm"],
        "Pipedrive":        ["pipedrive.com"],
        "Freshsales":       ["freshsales.io", "freshworks.com"],
        "ActiveCampaign":   ["activecampaign.com", "trackcmp.net", "activehosted.com"],
        "Marketo":          ["marketo.com", "mktoresp.com", "munchkin.marketo.net"],
        "Klaviyo":          ["klaviyo.com", "klaviyo.js"],
        "Mailchimp":        ["mailchimp.com", "chimpstatic.com", "list-manage.com"],
        "Brevo":            ["sendinblue.com", "brevo.com", "sibforms.com"],
        "Drip":             ["getdrip.com", "drip.com"],
        "Omnisend":         ["omnisend.com"],
        "Keap (Infusionsoft)": ["infusionsoft.com", "keap.com", "app.infusionsoft"],
    },

    "Analytics": {
        "Google Analytics": ["google-analytics.com", "googletagmanager.com", "gtag(", "ua-", "'g-"],
        "Google Tag Manager": ["googletagmanager.com", "gtm-", "gtm.js"],
        "Hotjar":           ["hotjar.com", "static.hotjar.com", "hjid", "hj("],
        "Mixpanel":         ["mixpanel.com", "mixpanel.init"],
        "Segment":          ["segment.com", "segment.io", "analytics.js", "cdn.segment"],
        "Amplitude":        ["amplitude.com", "amplitude.getInstance"],
        "Heap":             ["heap.io", "heap.js", "heap.load"],
        "Microsoft Clarity":["clarity.ms", "microsoft clarity"],
        "Plausible":        ["plausible.io", "data-domain"],
        "Matomo / Piwik":   ["matomo.js", "piwik.js", "matomo.php"],
        "Woopra":           ["woopra.com"],
        "Kissmetrics":      ["kissmetrics.com", "kissmetrics.js"],
    },

    "Live Chat / Support": {
        "Intercom":         ["intercom.io", "widget.intercom.io", "intercomSettings"],
        "Drift":            ["drift.com", "js.driftt.com", "drift.load"],
        "Zendesk":          ["zendesk.com", "zopim.com", "zopim("],
        "Tidio":            ["tidio.com", "tidio.co"],
        "Crisp":            ["crisp.chat", "client.crisp.chat"],
        "LiveChat":         ["livechatinc.com", "livechat.com"],
        "Tawk.to":          ["tawk.to", "tawkto"],
        "Freshchat":        ["freshchat.com", "freshdesk.com"],
        "Olark":            ["olark.com"],
        "HelpScout":        ["helpscout.net", "beacon-v2.helpscout.net"],
        "Chatwoot":         ["chatwoot.com"],
    },

    "Payments": {
        "Stripe":       ["stripe.com", "js.stripe.com", "stripe.js"],
        "PayPal":       ["paypal.com", "paypalobjects.com", "paypal.js"],
        "Square":       ["squareup.com", "square.com", "squaresandbox"],
        "Klarna":       ["klarna.com", "klarna.js"],
        "Braintree":    ["braintreepayments.com", "braintree-api.com"],
        "Razorpay":     ["razorpay.com", "checkout.razorpay"],
        "Paddle":       ["paddle.com", "cdn.paddle.com"],
        "Chargebee":    ["chargebee.com"],
        "Mollie":       ["mollie.com"],
        "Adyen":        ["adyen.com"],
    },

    "Advertising / Pixels": {
        "Facebook Pixel":   ["connect.facebook.net", "fbq(", "facebook-pixel"],
        "Google Ads":       ["googleadservices.com", "google_conversion", "goog_report"],
        "LinkedIn Insight": ["snap.licdn.com", "linkedin insight"],
        "TikTok Pixel":     ["analytics.tiktok.com", "tiktok pixel"],
        "Twitter/X Pixel":  ["static.ads-twitter.com", "twq("],
        "Pinterest Tag":    ["pintrk(", "ct.pinterest.com"],
        "Snapchat Pixel":   ["tr.snapchat.com", "snaptr("],
    },

    "Hosting / CDN": {
        "Cloudflare":   ["cf-ray", "__cfduid", "cloudflare"],
        "AWS":          ["x-amz-", "amazonaws.com", "aws-"],
        "Vercel":       ["x-vercel-", "vercel.app"],
        "Netlify":      ["x-nf-", "netlify.com", "netlify.app"],
        "Fastly":       ["x-fastly", "fastly"],
        "Akamai":       ["akamai", "akamaized.net"],
        "Google Cloud": ["x-cloud-trace-context", "googleusercontent.com"],
        "Azure":        ["x-ms-", "azurewebsites.net", "azure.com"],
    },

    "JavaScript Frameworks": {
        "React":        ["__react_devtools", "_reactrootcontainer", "react-root", "react.development"],
        "Vue.js":       ["__vue__", "vue.js", "vue.min.js", "data-v-"],
        "Angular":      ["ng-version", "angular.js", "angular.min.js", "ng-app"],
        "Next.js":      ["__next_data__", "_next/static", "next/dist"],
        "Nuxt.js":      ["__nuxt", "_nuxt/", "nuxt.js"],
        "jQuery":       ["jquery.js", "jquery.min.js", "jquery/"],
        "Alpine.js":    ["x-data=", "alpinejs"],
        "Svelte":       ["svelte", "__svelte"],
    },
}

# ─────────────────────────────────────────────────────────────────────────────

def normalize_url(url: str) -> str:
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def fetch_page(url: str, timeout: int = 12):
    try:
        resp = requests.get(url, headers=HEADERS_REQ, timeout=timeout, verify=False, allow_redirects=True)
        return resp
    except requests.exceptions.SSLError:
        try:
            url_http = url.replace("https://", "http://")
            return requests.get(url_http, headers=HEADERS_REQ, timeout=timeout, verify=False, allow_redirects=True)
        except Exception:
            return None
    except Exception:
        return None


def detect_technologies(url: str) -> dict:
    result = {"url": url, "status": "", "error": ""}
    for cat in FINGERPRINTS:
        result[cat] = []

    resp = fetch_page(normalize_url(url))
    if resp is None:
        result["status"] = "❌ Failed"
        result["error"] = "Could not connect"
        return result

    result["status"] = f"✅ {resp.status_code}"

    html = resp.text
    html_lower = html.lower()

    # Combine headers + cookies + html into one big searchable blob
    headers_str = " ".join(f"{k.lower()}: {v.lower()}" for k, v in resp.headers.items())
    cookies_str = " ".join(resp.cookies.keys()).lower()

    # Extract script srcs separately for cleaner matching
    soup = BeautifulSoup(html, "html.parser")
    scripts = " ".join(
        (tag.get("src", "") + " " + tag.string if tag.string else tag.get("src", ""))
        for tag in soup.find_all("script")
    ).lower()

    haystack = html_lower + " " + headers_str + " " + cookies_str + " " + scripts

    for category, techs in FINGERPRINTS.items():
        found = []
        for tech_name, patterns in techs.items():
            if any(p.lower() in haystack for p in patterns):
                found.append(tech_name)
        result[category] = found

    return result


def load_urls(filepath: str) -> list:
    with open(filepath, "r") as f:
        return [line.strip() for line in f if line.strip() and not line.startswith("#")]


def build_excel(results: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tech Stack Report"

    categories = list(FINGERPRINTS.keys())

    # ── Colour palette ─────────────────────────────────────────────────────
    HDR_BG   = "1F3864"   # dark navy
    HDR_FG   = "FFFFFF"
    CAT_COLORS = [
        "D9E1F2", "FCE4D6", "E2EFDA", "FFF2CC",
        "F4CCFF", "D0E4F7", "FFE0E0", "E8F5E9"
    ]
    YES_COLOR = "C6EFCE"   # green fill for detected cells
    NO_COLOR  = "FFCCCC"   # light red for nothing detected
    ALT_ROW   = "F7F9FC"

    hdr_font   = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    data_font  = Font(name="Arial", size=9)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Row 1: Main header ─────────────────────────────────────────────────
    total_cols = 3 + len(categories)   # URL | Status | Error | …categories
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    hdr_cell = ws.cell(row=1, column=1,
        value=f"🔍  Website Technology Report  —  Generated {datetime.now().strftime('%d %b %Y %H:%M')}")
    hdr_cell.font    = Font(name="Arial", bold=True, color=HDR_FG, size=13)
    hdr_cell.fill    = PatternFill("solid", start_color=HDR_BG)
    hdr_cell.alignment = center

    # ── Row 2: Column headers ──────────────────────────────────────────────
    col_headers = ["Website URL", "Status", "Error"] + categories
    for c, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cat_idx = c - 4   # 0-based for categories
        if c <= 3:
            bg = HDR_BG
        else:
            bg = "2E6099"   # slightly lighter navy for category headers
        cell.font      = Font(name="Arial", bold=True, color=HDR_FG, size=9)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = center
        cell.border    = thin_border

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, res in enumerate(results, start=3):
        row_bg = "FFFFFF" if row_idx % 2 == 0 else ALT_ROW

        # URL
        c = ws.cell(row=row_idx, column=1, value=res["url"])
        c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        c.alignment = left_wrap
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Status
        c = ws.cell(row=row_idx, column=2, value=res["status"])
        c.font = data_font
        c.alignment = center
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Error
        c = ws.cell(row=row_idx, column=3, value=res.get("error", ""))
        c.font = data_font
        c.alignment = left_wrap
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Category columns
        for cat_idx, cat in enumerate(categories, start=0):
            col = 4 + cat_idx
            detected = res.get(cat, [])
            value = ", ".join(detected) if detected else "—"
            cell_bg = CAT_COLORS[cat_idx % len(CAT_COLORS)] if detected else row_bg

            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = Font(name="Arial", size=9,
                          bold=bool(detected),
                          color="1A5E1A" if detected else "999999")
            c.fill = PatternFill("solid", start_color=cell_bg)
            c.alignment = left_wrap
            c.border = thin_border

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [38, 10, 20] + [24] * len(categories)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row heights ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32
    for r in range(3, 3 + len(results)):
        ws.row_dimensions[r].height = 38

    # ── Freeze top 2 rows ─────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40

    ws2.cell(row=1, column=1, value="Summary").font = Font(name="Arial", bold=True, size=12)
    ws2.cell(row=2, column=1, value="Total sites analysed").font = Font(name="Arial", size=10)
    ws2.cell(row=2, column=2, value=len(results)).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=3, column=1, value="Sites reachable").font = Font(name="Arial", size=10)
    ws2.cell(row=3, column=2, value=sum(1 for r in results if "✅" in r["status"])).font = Font(name="Arial", bold=True, size=10)

    ws2.cell(row=5, column=1, value="Technology").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=5, column=2, value="# Sites Using").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=5, column=3, value="Sites").font = Font(name="Arial", bold=True, size=10)

    tech_count: dict = {}
    for res in results:
        for cat in categories:
            for tech in res.get(cat, []):
                tech_count.setdefault(tech, []).append(res["url"])

    summary_row = 6
    for tech, sites in sorted(tech_count.items(), key=lambda x: -len(x[1])):
        ws2.cell(row=summary_row, column=1, value=tech).font = Font(name="Arial", size=9)
        ws2.cell(row=summary_row, column=2, value=len(sites)).font = Font(name="Arial", size=9)
        ws2.cell(row=summary_row, column=3, value=", ".join(sites)).font = Font(name="Arial", size=9)
        summary_row += 1

    wb.save(output_path)
    print(f"\n✅  Saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────

def main():
    urls_file  = sys.argv[1] if len(sys.argv) > 1 else "urls.txt"
    output     = sys.argv[2] if len(sys.argv) > 2 else f"tech_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        urls = load_urls(urls_file)
    except FileNotFoundError:
        print(f"❌  '{urls_file}' not found. Create it with one URL per line.")
        sys.exit(1)

    print(f"\n🔍  Analysing {len(urls)} URLs ...\n{'─'*60}")
    results = []

    for i, url in enumerate(urls, 1):
        print(f"  [{i}/{len(urls)}]  {url}", end="  ", flush=True)
        res = detect_technologies(url)
        results.append(res)
        found = sum(len(v) for k, v in res.items() if isinstance(v, list))
        print(f"→  {res['status']}  |  {found} tech detected")
        time.sleep(0.5)   # be polite

    build_excel(results, output)
    print(f"\n📊  Done — {len(results)} sites analysed\n")


if __name__ == "__main__":
    main()
