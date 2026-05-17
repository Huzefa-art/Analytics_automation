import requests
import re
import time
import sys
import random
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Suppress SSL warnings ──────────────────────────────────────────────────────
import urllib3
import asyncio
from ads_analyzer import check_meta_ads
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Email Regex
EMAIL_REGEX = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
]

def get_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://www.google.com/",
    }

# ─────────────────────────────────────────────────────────────────────────────
# FINGERPRINTS  { category: { tech_name: [patterns] } }
# patterns are checked against: html_lower, response headers, cookies, scripts
# ─────────────────────────────────────────────────────────────────────────────
FINGERPRINTS = {
    "CMS": {
        "WordPress":     ["wp-content", "wp-includes", "wp-json", "xmlrpc.php", "/wp-login"],
        "Shopify":       ["cdn.shopify.com", "shopify.theme", "myshopify.com", "shopify/assets"],
        "Wix":           ["wix.com/", "wixsite.com", "parastorage.com", "wixstatic.com"],
        "Squarespace":   ["squarespace.com", "static1.squarespace", "sqsp.net"],
        "Webflow":       ["webflow.com", "assets.website-files.com", "uploads-ssl.webflow"],
        "Drupal":        ["drupal.settings", "/sites/default/files", "drupal.js", "x-generator: drupal"],
        "Joomla":        ["/media/jui/", "joomla", "/components/com_"],
        "Magento":       ["mage.cookies", "/skin/frontend/", "magento", "mage/"],
        "BigCommerce":   ["cdn.bigcommerce.com", "bigcommerce", "stencil"],
        "Ghost":         ["ghost.io", "content/themes", "ghost/"],
        "Hubspot CMS":   ["hs-sites.com", "hubspotpreview.com"],
        "PrestaShop":    ["prestashop", "/themes/classic/", "presta_shop"],
        "OpenCart":      ["opencart", "catalog/view/theme"],
        "WooCommerce":   ["woocommerce", "wc-ajax", "wc_add_to_cart"],
        "Webnode":       ["webnode.com"],
        "Weebly":        ["weebly.com", "weeblycloud.com"],
        "Blogger":       ["blogger.com", "blogspot.com", "www.gstatic.com/blogger"],
    },

    "CRM / Marketing Automation": {
        "HubSpot":          ["hs-scripts.com", "hubspot.com", "_hsp", "hsforms.com", "hs-analytics"],
        "Salesforce":       ["salesforce.com", "pardot.com", "force.com", "exacttarget.com"],
        "Zoho CRM":         ["zoho.com", "zohopublic.com", "zohocrm"],
        "Pipedrive":        ["pipedrive.com"],
        "Freshsales":       ["freshsales.io", "freshworks.com"],
        "ActiveCampaign":   ["activecampaign.com", "trackcmp.net", "activehosted.com"],
        "Marketo":          ["marketo.com", "mktoresp.com", "munchkin.marketo.net"],
        "Klaviyo":          ["klaviyo.com", "klaviyo.js"],
        "Mailchimp":        ["mailchimp.com", "chimpstatic.com", "list-manage.com"],
        "Brevo":            ["sendinblue.com", "brevo.com", "sibforms.com"],
        "Drip":             ["getdrip.com", "drip.com"],
        "Omnisend":         ["omnisend.com"],
        "Keap (Infusionsoft)": ["infusionsoft.com", "keap.com", "app.infusionsoft"],
    },

    "Analytics": {
        "Google Analytics": ["google-analytics.com", "googletagmanager.com", "gtag(", "ua-", "'g-"],
        "Google Tag Manager": ["googletagmanager.com", "gtm-", "gtm.js"],
        "Hotjar":           ["hotjar.com", "static.hotjar.com", "hjid", "hj("],
        "Mixpanel":         ["mixpanel.com", "mixpanel.init"],
        "Segment":          ["segment.com", "segment.io", "analytics.js", "cdn.segment"],
        "Amplitude":        ["amplitude.com", "amplitude.getInstance"],
        "Heap":             ["heap.io", "heap.js", "heap.load"],
        "Microsoft Clarity":["clarity.ms", "microsoft clarity"],
        "Plausible":        ["plausible.io", "data-domain"],
        "Matomo / Piwik":   ["matomo.js", "piwik.js", "matomo.php"],
        "Woopra":           ["woopra.com"],
        "Kissmetrics":      ["kissmetrics.com", "kissmetrics.js"],
    },

    "Live Chat / Support": {
        "Intercom":         ["intercom.io", "widget.intercom.io", "intercomSettings"],
        "Drift":            ["drift.com", "js.driftt.com", "drift.load"],
        "Zendesk":          ["zendesk.com", "zopim.com", "zopim("],
        "Tidio":            ["tidio.com", "tidio.co"],
        "Crisp":            ["crisp.chat", "client.crisp.chat"],
        "LiveChat":         ["livechatinc.com", "livechat.com"],
        "Tawk.to":          ["tawk.to", "tawkto"],
        "Freshchat":        ["freshchat.com", "freshdesk.com"],
        "Olark":            ["olark.com"],
        "HelpScout":        ["helpscout.net", "beacon-v2.helpscout.net"],
        "Chatwoot":         ["chatwoot.com"],
    },

    "Payments": {
        "Stripe":       ["stripe.com", "js.stripe.com", "stripe.js"],
        "PayPal":       ["paypal.com", "paypalobjects.com", "paypal.js"],
        "Square":       ["squareup.com", "square.com", "squaresandbox"],
        "Klarna":       ["klarna.com", "klarna.js"],
        "Braintree":    ["braintreepayments.com", "braintree-api.com"],
        "Razorpay":     ["razorpay.com", "checkout.razorpay"],
        "Paddle":       ["paddle.com", "cdn.paddle.com"],
        "Chargebee":    ["chargebee.com"],
        "Mollie":       ["mollie.com"],
        "Adyen":        ["adyen.com"],
    },

    "Advertising / Pixels": {
        "Facebook Pixel":   ["connect.facebook.net", "fbq(", "facebook-pixel"],
        "Google Ads":       ["googleadservices.com", "google_conversion", "goog_report"],
        "LinkedIn Insight": ["snap.licdn.com", "linkedin insight"],
        "TikTok Pixel":     ["analytics.tiktok.com", "tiktok pixel"],
        "Twitter/X Pixel":  ["static.ads-twitter.com", "twq("],
        "Pinterest Tag":    ["pintrk(", "ct.pinterest.com"],
        "Snapchat Pixel":   ["tr.snapchat.com", "snaptr("],
    },

    "Hosting / CDN": {
        "Cloudflare":   ["cf-ray", "__cfduid", "cloudflare"],
        "AWS":          ["x-amz-", "amazonaws.com", "aws-"],
        "Vercel":       ["x-vercel-", "vercel.app"],
        "Netlify":      ["x-nf-", "netlify.com", "netlify.app"],
        "Fastly":       ["x-fastly", "fastly"],
        "Akamai":       ["akamai", "akamaized.net"],
        "Google Cloud": ["x-cloud-trace-context", "googleusercontent.com"],
        "Azure":        ["x-ms-", "azurewebsites.net", "azure.com"],
    },

    "JavaScript Frameworks": {
        "React":        ["__react_devtools", "_reactrootcontainer", "react-root", "react.development"],
        "Vue.js":       ["__vue__", "vue.js", "vue.min.js", "data-v-"],
        "Angular":      ["ng-version", "angular.js", "angular.min.js", "ng-app"],
        "Next.js":      ["__next_data__", "_next/static", "next/dist"],
        "Nuxt.js":      ["__nuxt", "_nuxt/", "nuxt.js"],
        "jQuery":       ["jquery.js", "jquery.min.js", "jquery/"],
        "Alpine.js":    ["x-data=", "alpinejs"],
        "Svelte":       ["svelte", "__svelte"],
    },
}

# ─────────────────────────────────────────────────────────────────────────────

def normalize_url(url: str) -> str:
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def fetch_page(url: str, timeout: int = 12):
    headers = get_headers()
    try:
        try:
            resp = requests.get(url, headers=headers, timeout=timeout, verify=False, allow_redirects=True)
            return resp
        except requests.exceptions.SSLError:
            url_http = url.replace("https://", "http://")
            return requests.get(url_http, headers=headers, timeout=timeout, verify=False, allow_redirects=True)
    except Exception:
        return None
def extract_social_links(soup) -> dict:
    fb_link = "N/A"
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        if "facebook.com/" in href and not any(x in href for x in ["sharer", "share.php", "messenger.com", "groups"]):
            fb_link = a["href"].split("?")[0].rstrip("/")
            break
    return {"Facebook": fb_link}

def extract_emails(soup, html_text) -> str:
    emails = set()
    
    # 1. Look for mailto: links
    for a in soup.select('a[href^="mailto:"]'):
        email = a["href"].replace("mailto:", "").split("?")[0].strip()
        if re.match(EMAIL_REGEX, email):
            emails.add(email)
            
    # 2. Look for patterns in text
    found_in_text = re.findall(EMAIL_REGEX, html_text)
    for email in found_in_text:
        # Avoid common false positives like image extensions
        if not email.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg')):
            emails.add(email)
            
    return ", ".join(sorted(list(emails))) if emails else "N/A"


def detect_technologies(url: str, no_tech: bool = False) -> dict:
    result = {"url": url, "status": "", "error": "", "facebook": "N/A", "email": "N/A"}
    for cat in FINGERPRINTS:
        result[cat] = []

    resp = fetch_page(normalize_url(url))
    if resp is None:
        result["status"] = "❌ Failed"
        result["error"] = "Could not connect"
        return result
    
    soup = BeautifulSoup(resp.text, "html.parser")
    socials = extract_social_links(soup)
    result["facebook"] = socials["Facebook"]
    result["email"] = extract_emails(soup, resp.text)

    result["status"] = f"✅ {resp.status_code}"

    if not no_tech:
        html = resp.text
        html_lower = html.lower()

        # Combine headers + cookies + html into one big searchable blob
        headers_str = " ".join(f"{k.lower()}: {v.lower()}" for k, v in resp.headers.items())
        cookies_str = " ".join(resp.cookies.keys()).lower()

        scripts = " ".join(
            (tag.get("src", "") + " " + tag.string if tag.string else tag.get("src", ""))
            for tag in soup.find_all("script")
        ).lower()

        haystack = html_lower + " " + headers_str + " " + cookies_str + " " + scripts

        for category, techs in FINGERPRINTS.items():
            found = []
            for tech_name, patterns in techs.items():
                if any(p.lower() in haystack for p in patterns):
                    found.append(tech_name)
            result[category] = found

    return result


def load_urls(filepath: str) -> list:
    with open(filepath, "r") as f:
        return [line.strip() for line in f if line.strip() and not line.startswith("#")]


def build_excel(results: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tech Stack Report"

    # ── SOCIAL & ADS SHEET ───────────────────────────────────────────────────
    ws_ads = wb.create_sheet("Facebook & Ads")

    categories = list(FINGERPRINTS.keys())

    # ── Colour palette ─────────────────────────────────────────────────────
    HDR_BG   = "1F3864"   # dark navy
    HDR_FG   = "FFFFFF"
    CAT_COLORS = [
        "D9E1F2", "FCE4D6", "E2EFDA", "FFF2CC",
        "F4CCFF", "D0E4F7", "FFE0E0", "E8F5E9"
    ]
    YES_COLOR = "C6EFCE"   # green fill for detected cells
    NO_COLOR  = "FFCCCC"   # light red for nothing detected
    ALT_ROW   = "F7F9FC"

    hdr_font   = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    data_font  = Font(name="Arial", size=9)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Row 1: Main header ─────────────────────────────────────────────────
    total_cols = 4 + len(categories)   # URL | Status | Error | Email | …categories
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    hdr_cell = ws.cell(row=1, column=1,
        value=f"🔍  Website Technology Report  —  Generated {datetime.now().strftime('%d %b %Y %H:%M')}")
    hdr_cell.font    = Font(name="Arial", bold=True, color=HDR_FG, size=13)
    hdr_cell.fill    = PatternFill("solid", start_color=HDR_BG)
    hdr_cell.alignment = center

    # ── Row 2: Column headers ──────────────────────────────────────────────
    col_headers = ["Website URL", "Status", "Error", "Email"] + categories
    for c, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cat_idx = c - 5   # 0-based for categories
        if c <= 4:
            bg = HDR_BG
        else:
            bg = "2E6099"   # slightly lighter navy for category headers
        cell.font      = Font(name="Arial", bold=True, color=HDR_FG, size=9)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = center
        cell.border    = thin_border

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, res in enumerate(results, start=3):
        row_bg = "FFFFFF" if row_idx % 2 == 0 else ALT_ROW

        # URL
        c = ws.cell(row=row_idx, column=1, value=res["url"])
        c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        c.alignment = left_wrap
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Status
        c = ws.cell(row=row_idx, column=2, value=res["status"])
        c.font = data_font
        c.alignment = center
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Error
        c = ws.cell(row=row_idx, column=3, value=res.get("error", ""))
        c.font = data_font
        c.alignment = left_wrap
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Email
        c = ws.cell(row=row_idx, column=4, value=res.get("email", "N/A"))
        c.font = data_font
        c.alignment = left_wrap
        c.fill = PatternFill("solid", start_color=row_bg)
        c.border = thin_border

        # Category columns
        for cat_idx, cat in enumerate(categories, start=0):
            col = 5 + cat_idx
            detected = res.get(cat, [])
            value = ", ".join(detected) if detected else "—"
            cell_bg = CAT_COLORS[cat_idx % len(CAT_COLORS)] if detected else row_bg

            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = Font(name="Arial", size=9,
                          bold=bool(detected),
                          color="1A5E1A" if detected else "999999")
            c.fill = PatternFill("solid", start_color=cell_bg)
            c.alignment = left_wrap
            c.border = thin_border

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [38, 10, 20, 25] + [24] * len(categories)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row heights ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32
    for r in range(3, 3 + len(results)):
        ws.row_dimensions[r].height = 38

    # ── Freeze top 2 rows ─────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── SOCIAL & ADS SHEET CONTENT ──────────────────────────────────────────
    ws_ads.column_dimensions["A"].width = 38
    ws_ads.column_dimensions["B"].width = 38
    ws_ads.column_dimensions["C"].width = 25
    ws_ads.column_dimensions["D"].width = 15
    ws_ads.column_dimensions["E"].width = 10
    ws_ads.column_dimensions["F"].width = 20

    ads_headers = ["Website URL", "Facebook Page", "Email", "Ads Active", "Ad Count", "Oldest Ad Date"]
    for c, h in enumerate(ads_headers, start=1):
        cell = ws_ads.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", start_color=HDR_BG)
        cell.alignment = center
        cell.border = thin_border

    for row_idx, res in enumerate(results, start=2):
        row_bg = "FFFFFF" if row_idx % 2 == 0 else ALT_ROW
        ws_ads.cell(row=row_idx, column=1, value=res["url"]).fill = PatternFill("solid", start_color=row_bg)
        ws_ads.cell(row=row_idx, column=2, value=res["facebook"]).fill = PatternFill("solid", start_color=row_bg)
        ws_ads.cell(row=row_idx, column=3, value=res.get("email", "N/A")).fill = PatternFill("solid", start_color=row_bg)
        
        ad_info = res.get("ad_info", {"active": "Checking...", "count": "—", "oldest_date": "—"})
        
        c4 = ws_ads.cell(row=row_idx, column=4, value=ad_info["active"])
        c5 = ws_ads.cell(row=row_idx, column=5, value=ad_info["count"])
        c6 = ws_ads.cell(row=row_idx, column=6, value=ad_info["oldest_date"])
        
        # Color coding for "Ads Active"
        if ad_info["active"] == "Yes":
           c4.font = Font(name="Arial", size=9, bold=True, color="1A5E1A")
           c4.fill = PatternFill("solid", start_color="C6EFCE")
        elif ad_info["active"] == "No":
           c4.font = Font(name="Arial", size=9, color="999999")
           c4.fill = PatternFill("solid", start_color="FFCCCC")
        else:
           c4.fill = PatternFill("solid", start_color=row_bg)

        c5.fill = PatternFill("solid", start_color=row_bg)
        c6.fill = PatternFill("solid", start_color=row_bg)
        
        for c in range(1, 7):
            ws_ads.cell(row=row_idx, column=c).border = thin_border
            if c != 4: ws_ads.cell(row=row_idx, column=c).font = data_font

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40

    ws2.cell(row=1, column=1, value="Summary").font = Font(name="Arial", bold=True, size=12)
    ws2.cell(row=2, column=1, value="Total sites analysed").font = Font(name="Arial", size=10)
    ws2.cell(row=2, column=2, value=len(results)).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=3, column=1, value="Sites reachable").font = Font(name="Arial", size=10)
    ws2.cell(row=3, column=2, value=sum(1 for r in results if "✅" in r["status"])).font = Font(name="Arial", bold=True, size=10)

    ws2.cell(row=5, column=1, value="Technology").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=5, column=2, value="# Sites Using").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=5, column=3, value="Sites").font = Font(name="Arial", bold=True, size=10)

    tech_count: dict = {}
    for res in results:
        for cat in categories:
            for tech in res.get(cat, []):
                tech_count.setdefault(tech, []).append(res["url"])

    summary_row = 6
    for tech, sites in sorted(tech_count.items(), key=lambda x: -len(x[1])):
        ws2.cell(row=summary_row, column=1, value=tech).font = Font(name="Arial", size=9)
        ws2.cell(row=summary_row, column=2, value=len(sites)).font = Font(name="Arial", size=9)
        ws2.cell(row=summary_row, column=3, value=", ".join(sites)).font = Font(name="Arial", size=9)
        summary_row += 1

    wb.save(output_path)
    print(f"\n✅  Saved: {output_path}")


def update_results_csv(tech_results):
    csv_path = "results.csv"
    import pandas as pd
    import os
    import re
    
    flat_results = []
    for r in tech_results:
        flat_r = {
            "Website": r["url"],
            "Facebook Page": r.get("facebook", "N/A"),
            "Email": r.get("email", "N/A"),
            "Ads Active": r.get("ad_info", {}).get("active", "N/A"),
            "Ad Count": r.get("ad_info", {}).get("count", "—"),
            "Oldest Ad Date": r.get("ad_info", {}).get("oldest_date", "—")
        }
        for cat in FINGERPRINTS:
            detected = r.get(cat, [])
            flat_r[cat] = ", ".join(detected) if detected else "N/A"
        flat_results.append(flat_r)
        
    df_tech = pd.DataFrame(flat_results)
    
    if os.path.exists(csv_path):
        try:
            df_orig = pd.read_csv(csv_path)
            
            def clean_url(u):
                if not isinstance(u, str):
                    return ""
                u = u.lower().strip()
                u = re.sub(r'^https?://', '', u)
                u = re.sub(r'^www\.', '', u)
                u = u.rstrip('/')
                return u
                
            df_orig['_match_url'] = df_orig['Website'].apply(clean_url)
            df_tech['_match_url'] = df_tech['Website'].apply(clean_url)
            
            # Remove any tech columns from df_orig before merge to avoid duplicates
            cols_to_drop = [c for c in df_tech.columns if c != '_match_url' and c in df_orig.columns]
            if cols_to_drop:
                df_orig = df_orig.drop(columns=cols_to_drop)
                
            df_merged = pd.merge(df_orig, df_tech, on='_match_url', how='left')
            df_merged = df_merged.drop(columns=['_match_url'])
            
            df_merged.to_csv(csv_path, index=False)
            print(f"Updated existing {csv_path} with detected technologies.")
        except Exception as e:
            print(f"Error merging with results.csv: {e}")
    else:
        df_new = pd.DataFrame(flat_results)
        df_new["Business Name"] = df_new["Website"]
        df_new["City"] = "N/A"
        df_new["Country"] = "N/A"
        df_new["Address"] = "N/A"
        df_new["Phone"] = "N/A"
        
        cols = ["Business Name", "Website", "City", "Country", "Address", "Phone", "Facebook Page", "Email", "Ads Active", "Ad Count", "Oldest Ad Date"] + list(FINGERPRINTS.keys())
        df_new = df_new.reindex(columns=cols)
        df_new.to_csv(csv_path, index=False)
        print(f"Created new {csv_path} with detected technologies.")


# ─────────────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("urls_file", nargs="?", default="urls.txt")
    parser.add_argument("output", nargs="?", default=None)
    parser.add_argument("--no-tech", action="store_true", help="Skip tech stack detection")
    parser.add_argument("--no-ads", action="store_true", help="Skip Meta Ads detection")
    args = parser.parse_args()

    urls_file = args.urls_file
    output = args.output if args.output else f"tech_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        urls = load_urls(urls_file)
    except FileNotFoundError:
        print(f"❌  '{urls_file}' not found. Create it with one URL per line.")
        sys.exit(1)

    print(f"\n🔍  Analysing {len(urls)} URLs ...\n{'─'*60}")
    results = []

    for i, url in enumerate(urls, 1):
        print(f"  [{i}/{len(urls)}]  {url}", end="  ", flush=True)
        res = detect_technologies(url, no_tech=args.no_tech)
        
        # Check Meta Ads if Facebook page found
        if not args.no_ads and res["facebook"] != "N/A":
            print(f"| Checking Ads ...", end=" ", flush=True)
            ad_info = asyncio.run(check_meta_ads(res["facebook"]))
            res["ad_info"] = ad_info
            print(f"→ {ad_info['active']}", end=" ", flush=True)
        else:
            res["ad_info"] = {"active": "N/A" if args.no_ads else "No", "count": 0, "oldest_date": "—"}

        results.append(res)
        found = sum(len(v) for k, v in res.items() if isinstance(v, list))
        print(f"| {found} tech detected")
        time.sleep(random.uniform(1.2, 3.5))   # Randomized delay to prevent blocking

    build_excel(results, output)
    print(f"\n📊  Done — {len(results)} sites analysed\n")
    
    try:
        update_results_csv(results)
    except Exception as e:
        print(f"Error updating results.csv: {e}")


if __name__ == "__main__":
    main()
